# core/trends.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Iterable
import re

from openpyxl import load_workbook


CATEGORY_LABELS = {
    "ACCA_LONGTERM": "ACCA LongTerm",
    "ACCA": "ACCA",
    "DCWAC": "DCwAC",
}


# Accept multiple possible header names for percent
PERCENT_HEADERS = {
    "PERCENT LOADING",
    "LIMVIOLPCT",
    "LIMITVIOLPCT",
    "LIM VIOL PCT",
    "PCT",
    "PERCENT",
}
ISSUE_HEADERS = {
    "RESULTING ISSUE",
    "RESULTINGISSUE",
    "ISSUE",
}
CONT_HEADERS = {
    "CONTINGENCY EVENTS",
    "CONTINGENCY",
    "CTGLABEL",
    "CONTINGENCY EVENT",
}
# Optional - not required for trends, but useful later
VALUE_HEADERS = {
    "CONTINGENCY VALUE",
    "CONTINGENCY VALUE (MVA)",
    "LIMVIOLVALUE",
    "VALUE",
}


def normalize_header(s: str) -> str:
    s = (s or "").strip()
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def normalize_issue_key(s: str) -> str:
    """
    Create a stable key for an issue so it matches across sheets.

    Keep it conservative (do NOT remove too much), just normalize whitespace.
    """
    s = (s or "").strip()
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def try_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # remove percent signs or stray chars
    s = s.replace("%", "").strip()
    try:
        return float(s)
    except Exception:
        return None


@dataclass
class IssueTrend:
    issue_key: str
    issue_display: str
    # sheet -> max percent in that sheet (None if missing)
    per_sheet: Dict[str, Optional[float]]

    @property
    def max_value(self) -> Optional[float]:
        vals = [v for v in self.per_sheet.values() if v is not None]
        return max(vals) if vals else None

    @property
    def count_present(self) -> int:
        return sum(1 for v in self.per_sheet.values() if v is not None)


@dataclass
class TrendResult:
    workbook_path: str
    category: str  # one of CATEGORY_LABELS keys
    sheet_order: List[str]
    issues: Dict[str, IssueTrend]  # issue_key -> IssueTrend


def _find_category_blocks(ws) -> Dict[str, List[Tuple[int, int]]]:
    """
    Returns: category_key -> list of (start_row, end_row) blocks.

    A block is detected by a row that contains the category label (e.g., "ACCA"),
    followed by a header row containing the needed columns, then data rows until
    next blank/next category label.
    """
    max_row = ws.max_row
    blocks: Dict[str, List[Tuple[int, int]]] = {k: [] for k in CATEGORY_LABELS.keys()}

    # Precompute quick label lookup
    label_to_key = {}
    for k, label in CATEGORY_LABELS.items():
        label_to_key[normalize_header(label)] = k

    # Identify rows where category header appears
    cat_rows: List[Tuple[int, str]] = []
    for r in range(1, max_row + 1):
        # Scan first ~10 columns to find category label
        found_key = None
        for c in range(1, min(11, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            hv = normalize_header(str(v))
            if hv in label_to_key:
                found_key = label_to_key[hv]
                break
        if found_key:
            cat_rows.append((r, found_key))

    # Add sentinel end
    cat_rows_sorted = sorted(cat_rows, key=lambda x: x[0])

    for idx, (r0, key) in enumerate(cat_rows_sorted):
        r_next = cat_rows_sorted[idx + 1][0] if idx + 1 < len(cat_rows_sorted) else max_row + 1
        # The block likely begins at r0 (category title row), headers on r0+1
        # But sometimes there can be an extra row; we’ll search forward a few rows for headers.
        header_row = None
        for rr in range(r0 + 1, min(r0 + 6, r_next)):
            # find columns in this row
            headers = [normalize_header(ws.cell(rr, cc).value if ws.cell(rr, cc).value is not None else "") for cc in range(1, ws.max_column + 1)]
            has_issue = any(h in ISSUE_HEADERS for h in headers)
            has_pct = any(h in PERCENT_HEADERS for h in headers)
            has_cont = any(h in CONT_HEADERS for h in headers)
            if has_issue and has_pct and has_cont:
                header_row = rr
                break
        if header_row is None:
            continue

        # Data starts next row
        data_start = header_row + 1
        data_end = data_start - 1
        for rr in range(data_start, r_next):
            # Stop if we hit another category label row early (defensive)
            hit_cat = False
            for cc in range(1, min(11, ws.max_column + 1)):
                v = ws.cell(rr, cc).value
                if v is None:
                    continue
                hv = normalize_header(str(v))
                if hv in label_to_key:
                    hit_cat = True
                    break
            if hit_cat:
                break

            # determine if row is "blank enough"
            row_vals = []
            for cc in range(1, min(ws.max_column, 12) + 1):
                row_vals.append(ws.cell(rr, cc).value)
            if all((v is None or str(v).strip() == "") for v in row_vals):
                # blank row ends the block (but allow trailing blanks)
                break

            data_end = rr

        if data_end >= data_start:
            blocks[key].append((header_row, data_end))

    return blocks


def _extract_sheet_category_max(ws, category_key: str) -> Dict[str, Tuple[str, float]]:
    """
    For one worksheet and one category, return:
      issue_key -> (issue_display, max_percent_in_sheet)
    Uses MAX per issue within the sheet.
    """
    blocks = _find_category_blocks(ws)
    cat_blocks = blocks.get(category_key, [])
    if not cat_blocks:
        return {}

    results: Dict[str, Tuple[str, float]] = {}

    for (header_row, end_row) in cat_blocks:
        # Map columns
        header_map: Dict[str, int] = {}
        for cc in range(1, ws.max_column + 1):
            hv = normalize_header(ws.cell(header_row, cc).value if ws.cell(header_row, cc).value is not None else "")
            if hv in ISSUE_HEADERS:
                header_map["issue"] = cc
            elif hv in PERCENT_HEADERS:
                header_map["pct"] = cc
            elif hv in CONT_HEADERS:
                header_map["cont"] = cc
            elif hv in VALUE_HEADERS:
                header_map["val"] = cc

        if "issue" not in header_map or "pct" not in header_map:
            continue

        issue_col = header_map["issue"]
        pct_col = header_map["pct"]

        for rr in range(header_row + 1, end_row + 1):
            issue_raw = ws.cell(rr, issue_col).value
            pct_raw = ws.cell(rr, pct_col).value
            if issue_raw is None:
                continue
            issue_display = str(issue_raw).strip()
            if not issue_display:
                continue
            pct = try_float(pct_raw)
            if pct is None:
                continue

            key = normalize_issue_key(issue_display)
            prev = results.get(key)
            if prev is None or pct > prev[1]:
                results[key] = (issue_display, pct)

    return results


def build_trends(
    workbook_path: str,
    category_key: str,
    min_percent: float = 80.0,
) -> TrendResult:
    """
    Build issue trends across all sheets for a category.
    Missing category in a sheet -> sheet contributes no issues (all None for those issues).
    """
    wb = load_workbook(workbook_path, data_only=True)
    sheet_names = wb.sheetnames[:]  # preserve order

    # First pass: collect per-sheet maxima
    per_sheet_issue: Dict[str, Dict[str, Tuple[str, float]]] = {}
    all_keys: Dict[str, str] = {}  # issue_key -> display

    for s in sheet_names:
        ws = wb[s]
        sheet_map = _extract_sheet_category_max(ws, category_key)
        per_sheet_issue[s] = sheet_map
        for k, (disp, _) in sheet_map.items():
            # keep first seen display
            all_keys.setdefault(k, disp)

    # Build IssueTrend objects
    issues: Dict[str, IssueTrend] = {}
    for issue_key, issue_disp in all_keys.items():
        per_sheet_vals: Dict[str, Optional[float]] = {}
        for s in sheet_names:
            entry = per_sheet_issue[s].get(issue_key)
            per_sheet_vals[s] = entry[1] if entry is not None else None
        it = IssueTrend(issue_key=issue_key, issue_display=issue_disp, per_sheet=per_sheet_vals)
        # Apply min filter based on MAX across all sheets
        mx = it.max_value
        if mx is not None and mx >= float(min_percent):
            issues[issue_key] = it

    return TrendResult(
        workbook_path=workbook_path,
        category=category_key,
        sheet_order=sheet_names,
        issues=issues,
    )