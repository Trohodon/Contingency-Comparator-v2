# core/batch_workbook_writer.py
#
# Batch workbook writer utilities (formatting + +/- outline grouping).

from __future__ import annotations

from typing import Set

import pandas as pd

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Styles (approximate the blue style from Combined_ViolationCTG_Comparison.xlsx)
HEADER_FILL = PatternFill("solid", fgColor="305496")  # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = HEADER_FILL
TITLE_FONT = Font(color="FFFFFF", bold=True, size=12)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CELL_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
CELL_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def sanitize_sheet_name(name: str) -> str:
    """
    Make a string safe to use as an Excel sheet name.
    """
    invalid = set(r'[]:*?/\\')
    cleaned = "".join(ch if ch not in invalid else "_" for ch in name)
    cleaned = cleaned.strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:31]


def apply_table_styles(ws: Worksheet):
    """
    Set reasonable column widths for a formatted comparison sheet.
    (No frozen panes – normal scrolling.)
    """
    widths = {
        2: 45,  # B: Contingency
        3: 45,  # C: Resulting Issue
        4: 15,  # D: Left %
        5: 15,  # E: Right %
        6: 22,  # F: Delta
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title_row(ws: Worksheet, row: int, title: str):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2)
    cell.value = title
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = CELL_ALIGN_CENTER


def _write_header_row(ws: Worksheet, row: int):
    headers = [
        "Contingency Events",
        "Resulting Issue",
        "Left %",
        "Right %",
        "Δ% (Right - Left) / Status",
    ]
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=2 + col_offset)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CELL_ALIGN_CENTER
        cell.border = THIN_BORDER


def _write_data_row(ws: Worksheet, row: int, cont: str, issue: str, left_pct, right_pct, delta: str):
    values = [cont, issue, left_pct, right_pct, delta]
    for col_offset, val in enumerate(values):
        cell = ws.cell(row=row, column=2 + col_offset)
        cell.value = val
        cell.border = THIN_BORDER

        if col_offset in (0, 1):  # text columns
            cell.alignment = CELL_ALIGN_WRAP
        else:
            cell.alignment = Alignment(horizontal="right", vertical="top")

        # number formatting for percentages
        if col_offset in (2, 3) and isinstance(val, (float, int)):
            cell.number_format = "0.00"


def _enable_outline(ws: Worksheet):
    # These properties help Excel show +/- groups nicely
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_properties.outlinePr.applyStyles = True


def write_formatted_pair_sheet(
    wb: Workbook,
    ws_name: str,
    df_pair: pd.DataFrame,
    expand_issues: bool = True,
):
    """
    Create one sheet in the batch workbook using the same blue-block style:

      Title row (merged B:F): 'ACCA LongTerm' / 'ACCA' / 'DCwAC'
      Header row
      Data rows

    If expand_issues=True:
      - group rows by ResultingIssue
      - keep the top row visible
      - hide the rest as outline level 1 (Excel +/- expand)
    """
    ws = wb.create_sheet(title=ws_name)
    apply_table_styles(ws)

    if df_pair.empty:
        ws.cell(row=2, column=2).value = "No rows above threshold."
        return

    if expand_issues:
        _enable_outline(ws)

    current_row = 2

    # These are the pretty names already in df_pair["CaseType"]
    for case_type_pretty in ["ACCA LongTerm", "ACCA", "DCwAC"]:
        sub = df_pair[df_pair["CaseType"] == case_type_pretty].copy()
        if sub.empty:
            continue

        # Title row
        _write_title_row(ws, current_row, case_type_pretty)
        current_row += 1

        # Header row
        _write_header_row(ws, current_row)
        current_row += 1

        if not expand_issues:
            # Write flat, no +/- grouping
            for _, r in sub.iterrows():
                _write_data_row(
                    ws,
                    current_row,
                    str(r.get("Contingency", "") or ""),
                    str(r.get("ResultingIssue", "") or ""),
                    r.get("LeftPct", None),
                    r.get("RightPct", None),
                    str(r.get("DeltaDisplay", "") or ""),
                )
                current_row += 1

            # Blank row between blocks
            current_row += 1
            continue

        # Expandable view:
        # group by ResultingIssue in the order already provided by comparator
        # (comparator sorts issue groups by max loading, then rows inside each group).
        for issue, g in sub.groupby("ResultingIssue", sort=False):
            g = g.copy()

            # Summary row = first row
            first = True
            summary_row_idx = None

            for _, r in g.iterrows():
                cont = str(r.get("Contingency", "") or "")
                issue_text = str(r.get("ResultingIssue", "") or "")
                left_pct = r.get("LeftPct", None)
                right_pct = r.get("RightPct", None)
                delta = str(r.get("DeltaDisplay", "") or "")

                _write_data_row(ws, current_row, cont, issue_text, left_pct, right_pct, delta)

                if first:
                    summary_row_idx = current_row
                    first = False
                else:
                    # detail rows: outline level 1 + hidden by default
                    ws.row_dimensions[current_row].outlineLevel = 1
                    ws.row_dimensions[current_row].hidden = True

                current_row += 1

            # Mark the summary row as "collapsed" so Excel shows the + button
            if summary_row_idx is not None:
                ws.row_dimensions[summary_row_idx].collapsed = True

        # Blank row between blocks (also helps separate outline regions visually)
        current_row += 1
