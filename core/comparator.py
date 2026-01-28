"""core/comparator.py

Parses the Combined_ViolationCTG_Comparison.xlsx workbook produced by the
contingency builder and generates:
  - Live view comparisons (DataFrame) for a single case type
  - Batch comparison workbook (many left/right pairs)

UPDATE (Limit column support):
  The source scenario sheets may be in either of these layouts:
    Old layout (5 cols):
      B Contingency Events | C Resulting Issue | D Contingency Value (MVA) | E Percent Loading
    New layout (6 cols):
      B Contingency Events | C Resulting Issue | D Limit | E Contingency Value (MVA) | F Percent Loading

  This module now detects header columns dynamically and carries LimViolLimit
  through comparisons so pair/batch sheets can show the Limit.
"""

from __future__ import annotations

import math
import os
from typing import Callable, Dict, List, Optional, Sequence, Tuple

import pandas as pd

try:
    from openpyxl import load_workbook, Workbook

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


# ----------------------------- helpers ----------------------------- #


def _norm_header(x) -> str:
    if x is None:
        return ""
    return str(x).strip().lower().replace("\n", " ")


def _detect_columns(ws, header_row: int) -> Dict[str, int]:
    """Return mapping of semantic column -> 1-based column index.

    We inspect header cells from B..H (2..8) and try to find:
      - ctg (contingency events / ctglabel)
      - issue (resulting issue / limviolid)
      - limit (limit)
      - value (contingency value / mva / value)
      - pct (percent loading / pct)
    """
    cols = {}
    for col_idx in range(2, 9):  # B..H
        v = ws.cell(row=header_row, column=col_idx).value
        h = _norm_header(v)
        if not h:
            continue

        if ("contingency" in h and "value" not in h) or "ctglabel" in h:
            cols["ctg"] = col_idx
            continue

        if "resulting issue" in h or "limviolid" in h or (("issue" in h) and ("resulting" in h)):
            cols["issue"] = col_idx
            continue

        # Limit column: avoid matching "value"
        if "limit" in h and "value" not in h:
            cols["limit"] = col_idx
            continue

        if "mva" in h or ("contingency value" in h) or (h == "value") or ("limviolvalue" in h):
            cols["value"] = col_idx
            continue

        if "percent" in h or "loading" in h or "pct" in h or "limviolpct" in h:
            cols["pct"] = col_idx
            continue

    return cols


def _to_float(x) -> float:
    if x is None:
        return float("nan")
    if isinstance(x, (int, float)):
        try:
            return float(x)
        except Exception:
            return float("nan")
    s = str(x).strip().replace("%", "")
    try:
        return float(s)
    except Exception:
        return float("nan")


def _is_nan(x) -> bool:
    return isinstance(x, float) and math.isnan(x)


def _choose_limit(left, right) -> str:
    """Choose a single display value for Limit.

    If both exist and match -> show one.
    If both exist and differ -> show "L:<left> | R:<right>".
    Else show whichever exists.
    """
    l = "" if left is None else str(left).strip()
    r = "" if right is None else str(right).strip()

    if l and r:
        # try numeric compare (many are numbers)
        lf = _to_float(l)
        rf = _to_float(r)
        if not _is_nan(lf) and not _is_nan(rf):
            if abs(lf - rf) < 1e-9:
                return f"{lf:g}"
            return f"L:{lf:g} | R:{rf:g}"
        if l == r:
            return l
        return f"L:{l} | R:{r}"

    return l or r or ""


# ----------------------------- public API ----------------------------- #


def list_sheets(workbook_path: str) -> List[str]:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required to read Excel workbooks in this build.")
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _parse_scenario_sheet(
    workbook_path: str, sheet_name: str, log_func: Optional[Callable[[str], None]] = None
) -> pd.DataFrame:
    """Parse one scenario sheet into a normalized DataFrame.

    Output columns:
      - CaseType (e.g., ACCA_LongTerm)
      - CTGLabel
      - LimViolID
      - LimViolLimit (may be blank if not present in source)
      - LimViolValue (may be blank if not present in source)
      - LimViolPct (string or numeric, as encountered)
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required to parse Excel workbooks in this build.")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        ws = wb[sheet_name]

        rows = []
        r = 1
        max_r = ws.max_row or 1

        while r <= max_r:
            cell_b = ws.cell(row=r, column=2).value  # B
            # Title rows are the case type names (single merged cell in B..F/G)
            title = str(cell_b).strip() if cell_b is not None else ""
            if title in ("ACCA LongTerm", "ACCA", "DCwAC"):
                # header row is next
                header_row = r + 1
                colmap = _detect_columns(ws, header_row)

                # Fallback for old layout if detection fails
                ctg_col = colmap.get("ctg", 2)
                issue_col = colmap.get("issue", 3)
                limit_col = colmap.get("limit")  # may be None
                value_col = colmap.get("value")
                pct_col = colmap.get("pct")

                # Old layout fallback: D=value, E=pct
                if value_col is None and pct_col is None:
                    value_col = 4
                    pct_col = 5

                # New layout fallback: D=limit, E=value, F=pct (common)
                if pct_col is None and value_col is not None and value_col == 5:
                    pct_col = 6

                # Start reading data after header row
                r = header_row + 1

                # Canonicalize title -> CaseType
                case_type = {
                    "ACCA LongTerm": "ACCA_LongTerm",
                    "ACCA": "ACCA_P1,2,4,7",
                    "DCwAC": "DCwACver_P1-7",
                }.get(title, title)

                while r <= max_r:
                    ctg = ws.cell(row=r, column=ctg_col).value
                    issue = ws.cell(row=r, column=issue_col).value

                    # stop on blank lines
                    if (ctg is None or str(ctg).strip() == "") and (issue is None or str(issue).strip() == ""):
                        break

                    lim = ws.cell(row=r, column=limit_col).value if limit_col else None
                    val = ws.cell(row=r, column=value_col).value if value_col else None
                    pct = ws.cell(row=r, column=pct_col).value if pct_col else None

                    issue_str = str(issue).strip() if issue is not None else ""
                    limviol_id = issue_str

                    rows.append(
                        {
                            "CaseType": case_type,
                            "CTGLabel": "" if ctg is None else str(ctg).strip(),
                            "LimViolID": limviol_id,
                            "LimViolLimit": "" if lim is None else lim,
                            "LimViolValue": "" if val is None else val,
                            "LimViolPct": "" if pct is None else pct,
                        }
                    )
                    r += 1

                # skip any blank separator rows after a block
                r += 1
                continue

            r += 1

        df = pd.DataFrame(rows)
        if df.empty:
            return df

        # Forward-fill within each case type where LimViolID is blank (detail rows)
        df["LimViolID"] = df["LimViolID"].replace("", pd.NA)
        df["LimViolID"] = df.groupby("CaseType")["LimViolID"].ffill()
        df["LimViolID"] = df["LimViolID"].fillna("")

        # Clean whitespace
        df["CTGLabel"] = df["CTGLabel"].astype(str).str.strip()
        df["LimViolID"] = df["LimViolID"].astype(str).str.strip()

        return df

    finally:
        wb.close()


def build_case_type_comparison(
    workbook_path: str,
    base_sheet: str,
    new_sheet: str,
    case_type: str,
    max_rows: Optional[int] = None,
    log_func: Optional[Callable[[str], None]] = None,
) -> pd.DataFrame:
    """Return a DataFrame comparing a single case type between two sheets.

    Output columns:
      Contingency, ResultingIssue, Limit, LeftPct, RightPct, DeltaPct
    """
    base_df = _parse_scenario_sheet(workbook_path, base_sheet, log_func)
    new_df = _parse_scenario_sheet(workbook_path, new_sheet, log_func)

    base_df = base_df[base_df["CaseType"] == case_type].copy()
    new_df = new_df[new_df["CaseType"] == case_type].copy()

    # Keys
    key_cols = ["CTGLabel", "LimViolID"]

    # Convert pct to float for comparisons
    base_df["_pct_num"] = base_df["LimViolPct"].apply(_to_float)
    new_df["_pct_num"] = new_df["LimViolPct"].apply(_to_float)

    # Keep worst per (CTGLabel, LimViolID)
    base_df = (
        base_df.sort_values(by=key_cols + ["_pct_num"], ascending=[True, True, False], kind="mergesort")
        .drop_duplicates(subset=key_cols, keep="first")
    )
    new_df = (
        new_df.sort_values(by=key_cols + ["_pct_num"], ascending=[True, True, False], kind="mergesort")
        .drop_duplicates(subset=key_cols, keep="first")
    )

    base_df = base_df.rename(columns={"_pct_num": "LeftPct", "LimViolLimit": "LeftLimit"})
    new_df = new_df.rename(columns={"_pct_num": "RightPct", "LimViolLimit": "RightLimit"})

    merged = pd.merge(
        base_df[key_cols + ["LeftPct", "LeftLimit"]],
        new_df[key_cols + ["RightPct", "RightLimit"]],
        on=key_cols,
        how="outer",
    )

    # Display columns
    merged["Contingency"] = merged["CTGLabel"].fillna("").astype(str)
    merged["ResultingIssue"] = merged["LimViolID"].fillna("").astype(str)
    merged["Limit"] = merged.apply(lambda r: _choose_limit(r.get("LeftLimit"), r.get("RightLimit")), axis=1)

    def _delta(row):
        l = row.get("LeftPct", float("nan"))
        r = row.get("RightPct", float("nan"))
        if _is_nan(l) or _is_nan(r):
            return float("nan")
        return r - l

    merged["DeltaPct"] = merged.apply(_delta, axis=1)

    # Sort by worst max pct
    merged["_max"] = merged.apply(
        lambda r: max([x for x in [r.get("LeftPct"), r.get("RightPct")] if not _is_nan(x)] or [float("nan")]), axis=1
    )
    merged = merged.sort_values(by=["_max", "Contingency", "ResultingIssue"], ascending=[False, True, True], na_position="last")
    merged = merged.drop(columns=["_max"])

    out = merged[["Contingency", "ResultingIssue", "Limit", "LeftPct", "RightPct", "DeltaPct"]].copy()

    if max_rows is not None and max_rows > 0:
        out = out.head(int(max_rows))

    return out


def build_pair_comparison_df(
    workbook_path: str,
    left_sheet: str,
    right_sheet: str,
    threshold: float,
    log_func: Optional[Callable[[str], None]] = None,
) -> pd.DataFrame:
    """Build the DataFrame used for a formatted pair sheet.

    Output columns:
      CaseType, Contingency, ResultingIssue, Limit, LeftPct, RightPct, DeltaDisplay
    """
    records = []
    for case_type in ("ACCA_LongTerm", "ACCA_P1,2,4,7", "DCwACver_P1-7"):
        df = build_case_type_comparison(
            workbook_path,
            base_sheet=left_sheet,
            new_sheet=right_sheet,
            case_type=case_type,
            max_rows=None,
            log_func=log_func,
        )
        if df.empty:
            continue

        for _, row in df.iterrows():
            left_pct = row.get("LeftPct", float("nan"))
            right_pct = row.get("RightPct", float("nan"))
            vals = [v for v in [left_pct, right_pct] if not _is_nan(v)]
            if not vals:
                continue
            if max(vals) < threshold:
                continue

            if _is_nan(left_pct) and not _is_nan(right_pct):
                delta_display = "Only in right"
            elif not _is_nan(left_pct) and _is_nan(right_pct):
                delta_display = "Only in left"
            elif not _is_nan(left_pct) and not _is_nan(right_pct):
                delta_display = f"{(right_pct - left_pct):.2f}"
            else:
                delta_display = ""

            records.append(
                {
                    "CaseType": case_type,
                    "Contingency": row.get("Contingency", ""),
                    "ResultingIssue": row.get("ResultingIssue", ""),
                    "Limit": row.get("Limit", ""),
                    "LeftPct": left_pct,
                    "RightPct": right_pct,
                    "DeltaDisplay": delta_display,
                }
            )

    return pd.DataFrame.from_records(records)


def build_batch_comparison_workbook(
    src_workbook: str,
    pairs: Sequence[Tuple[str, str]],
    threshold: float,
    output_path: str,
    log_func: Optional[Callable[[str], None]] = None,
    expandable_issue_view: bool = True,
) -> str:
    """Create a batch workbook with one formatted sheet per (left,right) pair."""
    if not pairs:
        raise ValueError("No comparison pairs provided.")

    # Import here to avoid circular imports
    from core.batch_sheet_writer import write_formatted_pair_sheet

    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required to build batch workbooks.")

    out_wb = Workbook()
    # Remove default
    if out_wb.active is not None:
        out_wb.remove(out_wb.active)

    for left_sheet, right_sheet in pairs:
        if log_func:
            log_func(f"Building pair: {left_sheet} vs {right_sheet}")

        df = build_pair_comparison_df(
            workbook_path=src_workbook,
            left_sheet=left_sheet,
            right_sheet=right_sheet,
            threshold=threshold,
            log_func=log_func,
        )

        safe_name = f"{left_sheet} vs {right_sheet}"
        safe_name = safe_name.replace("/", "-").replace("\\", "-")
        safe_name = safe_name[:31]

        ws = out_wb.create_sheet(title=safe_name)
        write_formatted_pair_sheet(
            ws=ws,
            df=df,
            left_label=left_sheet,
            right_label=right_sheet,
            expandable_issue_view=expandable_issue_view,
            log_func=log_func,
        )

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    out_wb.save(output_path)

    if log_func:
        log_func(f"Saved batch workbook: {output_path}")

    return output_path