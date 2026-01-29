# core/straight_comparison.py
#
# Builds and writes a "Straight Comparison" sheet that compares N original
# scenario sheets side-by-side (no pairwise deltas).
#
# Output is styled to match the blue-block look of the batch comparison sheets:
# - CaseType blocks (ACCA LongTerm / ACCA / DCwAC)
# - Columns: Contingency Events | Resulting Issue | Limit | <one % column per scenario>
# - Optional Excel +/- outline grouped by Resulting Issue (summary row ABOVE details)
# - The top (max) row per Resulting Issue is bolded (like the batch sheets)
#
# UPDATED (Limit column):
# - Parses both old formatted sheets (no Limit) and new ones (with Limit).
# - Writes "Limit" in column D, and scenario % columns start at column E.

from __future__ import annotations

from typing import Dict, List, Sequence, Tuple, Optional
import math
import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# CaseType mappings (same as comparator.py)
# ---------------------------------------------------------------------------

CANONICAL_CASE_TYPES = {
    "ACCA LongTerm": "ACCA_LongTerm",
    "ACCA Long Term": "ACCA_LongTerm",
    "ACCA": "ACCA_P1,2,4,7",
    "DCwAC": "DCwACver_P1-7",
}

CASE_TYPES_CANONICAL: List[str] = [
    "ACCA_LongTerm",
    "ACCA_P1,2,4,7",
    "DCwACver_P1-7",
]

CANONICAL_TO_PRETTY = {
    "ACCA_LongTerm": "ACCA LongTerm",
    "ACCA_P1,2,4,7": "ACCA",
    "DCwACver_P1-7": "DCwAC",
}

_PRETTY_CASE_TITLES = ("ACCA LongTerm", "ACCA Long Term", "ACCA", "DCwAC")


# ---------------------------------------------------------------------------
# Parsing the formatted scenario sheets (FAST iter_rows version)
# ---------------------------------------------------------------------------

def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _header_has_limit_from_row(d_val) -> bool:
    if isinstance(d_val, str) and "limit" in d_val.strip().lower():
        return True
    return False


def _parse_scenario_sheet(ws: Worksheet, log_func=None) -> pd.DataFrame:
    """
    Parse one formatted scenario sheet into rows with:
        CaseType, CTGLabel, LimViolID, LimViolLimit, LimViolPct

    Supports:
      - OLD format:  B=CTGLabel, C=LimViolID, D=LimViolValue, E=LimViolPct
      - NEW format:  B=CTGLabel, C=LimViolID, D=LimViolLimit, E=LimViolValue, F=LimViolPct
    """
    records: List[Dict] = []

    current_case_type = None
    skip_rows = 0
    last_issue = None
    has_limit: Optional[bool] = None

    # Read columns B..F (2..6) so we can support both formats.
    for (b, c, d, e, f) in ws.iter_rows(min_row=1, min_col=2, max_col=6, values_only=True):
        # Not currently inside a case-type block
        if current_case_type is None:
            if isinstance(b, str) and b.strip():
                pretty = b.strip()
                current_case_type = CANONICAL_CASE_TYPES.get(pretty, pretty)
                skip_rows = 1  # next row is header
                last_issue = None
                has_limit = None
            continue

        # Inside a block: header row
        if skip_rows > 0:
            # Determine format by inspecting column D header
            has_limit = _header_has_limit_from_row(d)
            skip_rows -= 1
            continue

        # End of block condition: blank line across expected columns
        if has_limit:
            # NEW: B,C,D,E,F must all be blank to end
            if _is_blank(b) and _is_blank(c) and _is_blank(d) and _is_blank(e) and _is_blank(f):
                current_case_type = None
                skip_rows = 0
                last_issue = None
                has_limit = None
                continue
        else:
            # OLD: B,C,D,E blank to end
            if _is_blank(b) and _is_blank(c) and _is_blank(d) and _is_blank(e):
                current_case_type = None
                skip_rows = 0
                last_issue = None
                has_limit = None
                continue

        # Forward fill issue if missing
        if _is_blank(c) and last_issue is not None:
            c = last_issue
        else:
            if not _is_blank(c):
                last_issue = c

        if has_limit:
            lim = d
            pct = f
        else:
            lim = None
            pct = e

        records.append(
            {
                "CaseType": current_case_type,
                "CTGLabel": b,
                "LimViolID": c,
                "LimViolLimit": lim,
                "LimViolPct": pct,
            }
        )

    df = pd.DataFrame.from_records(records)
    if log_func:
        log_func(f"Parsed {len(df)} rows from sheet '{ws.title}' for straight comparison.")
    return df


def discover_scenario_sheets(workbook_path: str, log_func=None, scan_rows: int = 300) -> List[str]:
    """
    FAST scenario sheet discovery:
    A "scenario sheet" is any sheet that *looks* like the formatted output
    (i.e., contains at least one of the case-type titles in column B early on).
    """
    if not os.path.isfile(workbook_path):
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    found: List[str] = []
    for name in wb.sheetnames:
        try:
            ws = wb[name]
            hit = False
            for (b_val,) in ws.iter_rows(min_row=1, max_row=scan_rows, min_col=2, max_col=2, values_only=True):
                if isinstance(b_val, str):
                    s = b_val.strip()
                    if s in _PRETTY_CASE_TITLES:
                        hit = True
                        break
            if hit:
                found.append(name)
        except Exception:
            continue

    if log_func:
        log_func(f"Discovered {len(found)} scenario sheets for Straight Comparison (fast scan).")
    return found


# ---------------------------------------------------------------------------
# Build a straight comparison dataframe (loads workbook ONCE)
# ---------------------------------------------------------------------------

def _safe_float(x):
    try:
        if x is None:
            return None
        if isinstance(x, float) and math.isnan(x):
            return None
        return float(x)
    except Exception:
        return None


def build_straight_comparison_df(
    workbook_path: str,
    sheet_names: Sequence[str],
    threshold: float,
    log_func=None,
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Returns (df, ordered_case_labels).

    df columns:
      CaseType, Contingency, ResultingIssue, Limit, <case_label_1>, <case_label_2>, ...

    Threshold:
      keep rows where max across cases >= threshold.
    """
    if not sheet_names:
        return pd.DataFrame(columns=["CaseType", "Contingency", "ResultingIssue", "Limit"]), []

    # Create short-but-unique labels for column headers
    used: set[str] = set()
    labels: List[str] = []

    def make_label(name: str) -> str:
        base = str(name).strip()
        lab = base if len(base) <= 18 else base[:18]
        candidate = lab
        k = 2
        while candidate in used or candidate == "":
            suffix = f"_{k}"
            candidate = (lab[: max(1, 18 - len(suffix))] + suffix)
            k += 1
        used.add(candidate)
        return candidate

    for s in sheet_names:
        labels.append(make_label(s))

    # Load workbook ONCE (huge speedup)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    master: Optional[pd.DataFrame] = None

    for sheet_name, col_label in zip(sheet_names, labels):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        df = _parse_scenario_sheet(ws, log_func=log_func)
        if df.empty:
            continue

        df["CaseTypePretty"] = df["CaseType"].map(CANONICAL_TO_PRETTY).fillna(df["CaseType"])
        df = df.rename(columns={"CTGLabel": "Contingency", "LimViolID": "ResultingIssue"})
        df[col_label] = df["LimViolPct"].apply(_safe_float)

        # Keep limit as a display value (can be numeric or text)
        df["Limit"] = df.get("LimViolLimit")

        # Collapse duplicates by max pct within this scenario.
        # For Limit, take the first non-blank value encountered.
        def _first_nonblank(s: pd.Series):
            for v in s.tolist():
                if v is None:
                    continue
                if isinstance(v, float) and math.isnan(v):
                    continue
                if isinstance(v, str) and v.strip() == "":
                    continue
                return v
            return None

        df = (
            df.groupby(["CaseTypePretty", "Contingency", "ResultingIssue"], as_index=False)
              .agg({"Limit": _first_nonblank, col_label: "max"})
        )

        if master is None:
            master = df
        else:
            merged = pd.merge(
                master,
                df,
                on=["CaseTypePretty", "Contingency", "ResultingIssue"],
                how="outer",
                suffixes=("", "__new"),
            )
            # Combine limits (prefer existing, else new)
            if "Limit__new" in merged.columns:
                merged["Limit"] = merged.get("Limit").combine_first(merged.get("Limit__new"))
                merged = merged.drop(columns=["Limit__new"])
            master = merged

    if master is None or master.empty:
        out = pd.DataFrame(columns=["CaseType", "Contingency", "ResultingIssue", "Limit"] + labels)
        return out, labels

    master = master.rename(columns={"CaseTypePretty": "CaseType"})

    case_cols = [c for c in labels if c in master.columns]

    if case_cols:
        max_series = master[case_cols].max(axis=1, skipna=True)
        master = master[max_series.fillna(float("-inf")) >= float(threshold)].copy()

        master["_SortKey"] = max_series
        master = master.sort_values(
            by=["CaseType", "_SortKey"], ascending=[True, False], na_position="last"
        ).drop(columns=["_SortKey"])

    # Ensure every requested label exists
    for c in labels:
        if c not in master.columns:
            master[c] = None

    master = master[["CaseType", "Contingency", "ResultingIssue", "Limit"] + labels].copy()
    return master, labels


# ---------------------------------------------------------------------------
# Worksheet writer (blue-block style + optional +/-)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = HEADER_FILL
TITLE_FONT = Font(color="FFFFFF", bold=True, size=12)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CELL_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
CELL_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_table_styles(ws: Worksheet, num_cases: int):
    ws.column_dimensions[get_column_letter(2)].width = 45  # B: Contingency
    ws.column_dimensions[get_column_letter(3)].width = 45  # C: Issue
    ws.column_dimensions[get_column_letter(4)].width = 15  # D: Limit

    # Scenario % columns start at E now
    for i in range(num_cases):
        ws.column_dimensions[get_column_letter(5 + i)].width = 12

    try:
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
        ws.sheet_properties.outlinePr.applyStyles = True
    except Exception:
        pass

    try:
        ws.sheet_view.showOutlineSymbols = True
    except Exception:
        pass


def _write_title_row(ws: Worksheet, row: int, title: str, last_col: int):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=2)
    cell.value = title
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = CELL_ALIGN_CENTER


def _write_header_row(ws: Worksheet, row: int, case_labels: Sequence[str]):
    headers = ["Contingency Events", "Resulting Issue", "Limit"] + list(case_labels)
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=2 + col_offset)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CELL_ALIGN_CENTER
        cell.border = THIN_BORDER


def _write_row(
    ws: Worksheet,
    row: int,
    values: Sequence,
    *,
    outline_level: int = 0,
    hidden: bool = False,
    bold: bool = False,
):
    for col_offset, val in enumerate(values):
        cell = ws.cell(row=row, column=2 + col_offset)
        cell.value = val
        cell.border = THIN_BORDER

        if bold:
            base = cell.font or Font()
            cell.font = Font(
                name=base.name,
                size=base.size,
                bold=True,
                italic=base.italic,
                vertAlign=base.vertAlign,
                underline=base.underline,
                strike=base.strike,
                color=base.color,
            )

        if col_offset in (0, 1):
            cell.alignment = CELL_ALIGN_WRAP
        else:
            cell.alignment = Alignment(horizontal="right", vertical="top")

        # Case % columns are now offsets >= 3
        if col_offset >= 3 and isinstance(val, (float, int)):
            cell.number_format = "0.00"

    try:
        ws.row_dimensions[row].outlineLevel = int(outline_level)
        ws.row_dimensions[row].hidden = bool(hidden)
    except Exception:
        pass


def _max_across_cases(row: pd.Series, case_cols: Sequence[str]) -> float:
    vals = []
    for c in case_cols:
        v = row.get(c, None)
        if v is None:
            continue
        if isinstance(v, float) and math.isnan(v):
            continue
        try:
            vals.append(float(v))
        except Exception:
            continue
    return max(vals) if vals else float("-inf")


def write_formatted_straight_sheet(
    wb: Workbook,
    ws_name: str,
    df: pd.DataFrame,
    case_labels: Sequence[str],
    *,
    expandable_issue_view: bool = True,
):
    ws = wb.create_sheet(title=ws_name)
    _apply_table_styles(ws, num_cases=len(case_labels))

    if df is None or df.empty:
        ws.cell(row=2, column=2).value = "No rows above threshold."
        return

    current_row = 2
    case_cols = list(case_labels)

    # Columns: B..(D + num_cases)  => last_col = 4 + num_cases
    last_col = 4 + len(case_cols)

    for case_type_pretty in ["ACCA LongTerm", "ACCA", "DCwAC"]:
        sub = df[df["CaseType"] == case_type_pretty].copy()
        if sub.empty:
            continue

        _write_title_row(ws, current_row, case_type_pretty, last_col=last_col)
        current_row += 1
        _write_header_row(ws, current_row, case_cols)
        current_row += 1

        if not expandable_issue_view:
            for _, r in sub.iterrows():
                cont = str(r.get("Contingency", "") or "")
                issue = str(r.get("ResultingIssue", "") or "")
                limit = r.get("Limit", None)
                vals = [cont, issue, limit] + [r.get(c, None) for c in case_cols]
                _write_row(ws, current_row, vals)
                current_row += 1
            current_row += 1
            continue

        sub["_SortKey"] = sub.apply(lambda rr: _max_across_cases(rr, case_cols), axis=1)
        group_max = sub.groupby("ResultingIssue")["_SortKey"].max().sort_values(ascending=False)
        ordered_issues = list(group_max.index)

        for issue_key in ordered_issues:
            g = sub[sub["ResultingIssue"] == issue_key].copy()
            if g.empty:
                continue

            g = g.sort_values(by="_SortKey", ascending=False, na_position="last")

            summary_row_index = None
            first = True

            for _, r in g.iterrows():
                cont = str(r.get("Contingency", "") or "")
                issue = str(r.get("ResultingIssue", "") or "")
                issue_display = issue if first else ""
                limit = r.get("Limit", None)

                vals = [cont, issue_display, limit] + [r.get(c, None) for c in case_cols]

                if first:
                    summary_row_index = current_row

                _write_row(
                    ws,
                    current_row,
                    vals,
                    outline_level=0 if first else 1,
                    hidden=False if first else True,
                    bold=True if first else False,
                )
                current_row += 1
                first = False

            try:
                if summary_row_index is not None and len(g) > 1:
                    ws.row_dimensions[summary_row_index].collapsed = True
            except Exception:
                pass

        current_row += 1