# core/batch_sheet_writer.py
#
# Writes a batch comparison worksheet using the same blue-block style
# as the Combined_ViolationCTG_Comparison workbook.
#
# Key behavior:
# - Expandable issue view uses Excel outline (+/-)
# - Summary row is ABOVE detail rows (so +/- appears at the TOP like your "left" workbook)
# - The first (max) row per Resulting Issue is bolded to stand out
#
# UPDATED (Limit column):
# - Adds a "Limit" column (associated with Resulting Issue) in column D.
# - Output columns now:
#     B Contingency Events | C Resulting Issue | D Limit | E Left % | F Right % | G Δ% / Status

from __future__ import annotations

from typing import Optional
import math
import pandas as pd

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===== Formatting helpers ====================================================

def apply_table_styles(ws: Worksheet):
    """Set reasonable column widths and outline settings for a formatted comparison sheet."""
    widths = {
        2: 45,  # B: Contingency Events
        3: 45,  # C: Resulting Issue
        4: 15,  # D: Limit
        5: 15,  # E: Left %
        6: 15,  # F: Right %
        7: 22,  # G: Delta / Status
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Make outline symbols visible + summary row above details
    try:
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
        ws.sheet_properties.outlinePr.applyStyles = True
    except Exception:
        pass

    try:
        ws.sheet_view.showOutlineSymbols = True
    except Exception:
        pass


HEADER_FILL = PatternFill("solid", fgColor="305496")  # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = HEADER_FILL
TITLE_FONT = Font(color="FFFFFF", bold=True, size=12)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CELL_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
CELL_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _is_nan(x) -> bool:
    return isinstance(x, float) and math.isnan(x)


def _max_pct(left: Optional[float], right: Optional[float]) -> float:
    vals = []
    if left is not None and not _is_nan(left):
        vals.append(float(left))
    if right is not None and not _is_nan(right):
        vals.append(float(right))
    return max(vals) if vals else float("-inf")


def _normalize_issue_series(series: pd.Series) -> pd.Series:
    """Forward-fill blanks so blank ResultingIssue inherits the issue above."""
    s = series.copy()

    def is_blank(v) -> bool:
        if v is None:
            return True
        if isinstance(v, float) and math.isnan(v):
            return True
        if isinstance(v, str) and v.strip() == "":
            return True
        return False

    mask = s.apply(is_blank)
    s = s.mask(mask)
    s = s.ffill()
    return s.fillna("")  # final safety


def _write_title_row(ws: Worksheet, row: int, title: str):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)  # B..G
    cell = ws.cell(row=row, column=2)
    cell.value = title
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = CELL_ALIGN_CENTER


def _write_header_row(ws: Worksheet, row: int):
    headers = [
        "Contingency Events",
        "Resulting Issue",
        "Limit",
        "Left %",
        "Right %",
        "Δ% (Right - Left) / Status",
    ]
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=2 + col_offset)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CELL_ALIGN_CENTER
        cell.border = THIN_BORDER


def _write_data_row(
    ws: Worksheet,
    row: int,
    cont: str,
    issue: str,
    limit,
    left_pct,
    right_pct,
    delta,
    *,
    outline_level: int = 0,
    hidden: bool = False,
    bold: bool = False,
):
    values = [cont, issue, limit, left_pct, right_pct, delta]

    for col_offset, val in enumerate(values):
        cell = ws.cell(row=row, column=2 + col_offset)
        cell.value = val
        cell.border = THIN_BORDER

        if bold:
            base = cell.font or Font()
            cell.font = Font(
                name=base.name,
                size=base.size,
                bold=True,
                italic=base.italic,
                vertAlign=base.vertAlign,
                underline=base.underline,
                strike=base.strike,
                color=base.color,
            )

        # Align
        if col_offset in (0, 1):
            cell.alignment = CELL_ALIGN_WRAP
        else:
            cell.alignment = Alignment(horizontal="right", vertical="top")

        # Number formats
        # Left/Right are now offsets 3 and 4
        if col_offset in (3, 4) and isinstance(val, (float, int)):
            cell.number_format = "0.00"

    # Outline / hidden controls (Excel +/-)
    try:
        ws.row_dimensions[row].outlineLevel = int(outline_level)
        ws.row_dimensions[row].hidden = bool(hidden)
    except Exception:
        pass


def write_formatted_pair_sheet(
    wb: Workbook,
    ws_name: str,
    df_pair: pd.DataFrame,
    *,
    expandable_issue_view: bool = True,
):
    """Create one sheet in the batch workbook using the blue-block style."""
    ws = wb.create_sheet(title=ws_name)
    apply_table_styles(ws)

    if df_pair is None or df_pair.empty:
        ws.cell(row=2, column=2).value = "No rows above threshold."
        return

    current_row = 2

    for case_type_pretty in ["ACCA LongTerm", "ACCA", "DCwAC"]:
        sub = df_pair[df_pair["CaseType"] == case_type_pretty].copy()
        if sub.empty:
            continue

        # Normalize blank issues -> same as above (safety net)
        if "ResultingIssue" not in sub.columns:
            sub["ResultingIssue"] = ""
        sub["ResultingIssue"] = _normalize_issue_series(sub["ResultingIssue"])

        # Title + header rows
        _write_title_row(ws, current_row, case_type_pretty)
        current_row += 1
        _write_header_row(ws, current_row)
        current_row += 1

        if not expandable_issue_view:
            for _, r in sub.iterrows():
                cont = str(r.get("Contingency", "") or "")
                issue = str(r.get("ResultingIssue", "") or "")
                limit = r.get("Limit", None)
                left_pct = r.get("LeftPct", None)
                right_pct = r.get("RightPct", None)
                delta = str(r.get("DeltaDisplay", "") or "")

                _write_data_row(
                    ws,
                    current_row,
                    cont,
                    issue,
                    limit,
                    left_pct,
                    right_pct,
                    delta,
                    outline_level=0,
                    hidden=False,
                    bold=False,
                )
                current_row += 1

            current_row += 1
            continue

        # Expandable: group by issue, sort each group by max pct desc
        sub["_SortKey"] = sub.apply(
            lambda r: _max_pct(r.get("LeftPct", None), r.get("RightPct", None)),
            axis=1,
        )

        group_max = sub.groupby("ResultingIssue")["_SortKey"].max().sort_values(ascending=False)
        ordered_issues = list(group_max.index)

        for issue_key in ordered_issues:
            g = sub[sub["ResultingIssue"] == issue_key].copy()
            if g.empty:
                continue

            g = g.sort_values(by="_SortKey", ascending=False, na_position="last")

            summary_row_index = None
            first = True

            for _, r in g.iterrows():
                cont = str(r.get("Contingency", "") or "")
                issue = str(r.get("ResultingIssue", "") or "")

                # Blank issue text for hidden detail rows (readability)
                issue_display = issue if first else ""

                limit = r.get("Limit", None)
                left_pct = r.get("LeftPct", None)
                right_pct = r.get("RightPct", None)
                delta = str(r.get("DeltaDisplay", "") or "")

                if first:
                    summary_row_index = current_row

                _write_data_row(
                    ws,
                    current_row,
                    cont,
                    issue_display,
                    limit,
                    left_pct,
                    right_pct,
                    delta,
                    outline_level=0 if first else 1,
                    hidden=False if first else True,
                    bold=True if first else False,
                )
                current_row += 1
                first = False

            # Mark the summary row as collapsed when there are details
            try:
                if summary_row_index is not None and len(g) > 1:
                    ws.row_dimensions[summary_row_index].collapsed = True
            except Exception:
                pass

        # Blank row between blocks
        current_row += 1