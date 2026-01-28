import os
import pandas as pd

from .case_finder import TARGET_PATTERNS

# Try to import openpyxl for formatting + outline grouping
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Pretty display names for the three case types
PRETTY_CASE_NAMES = {
    "ACCA_LongTerm": "ACCA LongTerm",
    "ACCA_P1,2,4,7": "ACCA",
    "DCwACver_P1-7": "DCwAC",
}


def _to_float_series(series: pd.Series) -> pd.Series:
    """Convert a LimViolPct-like series to float safely."""
    if series is None:
        return pd.Series(dtype="float64")
    if pd.api.types.is_numeric_dtype(series):
        return series.astype(float)
    cleaned = series.astype(str).str.replace("%", "", regex=False).str.strip()
    return pd.to_numeric(cleaned, errors="coerce")


def _build_simple_workbook(root_folder, folder_to_case_csvs, log_func=None):
    """
    Fallback: simple one-sheet-per-scenario workbook, no fancy formatting,
    and no outline dropdown grouping.
    """
    if not folder_to_case_csvs:
        if log_func:
            log_func("No data to build combined workbook.")
        return None

    workbook_path = os.path.join(root_folder, "Combined_ViolationCTG_Comparison.xlsx")

    if log_func:
        log_func(f"\nBuilding SIMPLE combined workbook:\n  {workbook_path}")

    try:
        writer = pd.ExcelWriter(workbook_path)
    except Exception as e:
        if log_func:
            log_func(f"ERROR: Could not create ExcelWriter: {e}")
        return None

    try:
        with writer as xls_writer:
            for folder_name, case_map in folder_to_case_csvs.items():
                dfs = []
                for label in TARGET_PATTERNS:
                    csv_path = case_map.get(label)
                    if not csv_path:
                        continue
                    try:
                        df = pd.read_csv(csv_path)
                        df.insert(0, "CaseType", label)
                        dfs.append(df)
                    except Exception as e:
                        if log_func:
                            log_func(f"  [{folder_name}] WARNING: Failed to read {csv_path}: {e}")

                if not dfs:
                    continue

                combined = pd.concat(dfs, ignore_index=True)
                sheet_name = (folder_name or "Sheet").strip()[:31] or "Sheet"
                combined.to_excel(xls_writer, sheet_name=sheet_name, index=False)

    except Exception as e:
        if log_func:
            log_func(f"ERROR while building simple workbook: {e}")
        return None

    if log_func:
        log_func("Simple combined workbook build complete.")
    return workbook_path


def build_workbook(root_folder, folder_to_case_csvs, group_details: bool = True, log_func=None):
    """
    Build a combined Excel workbook with one sheet per subfolder.

    group_details=True:
        - Within each CaseType block, group rows by LimViolID.
        - Show the highest LimViolPct row per LimViolID.
        - Collapse (hide) the other contingencies under an Excel outline dropdown.
        - Sort the groups so the WORST (highest percent loading) issues appear first.
    """
    if not folder_to_case_csvs:
        if log_func:
            log_func("No data to build combined workbook.")
        return None

    if not OPENPYXL_AVAILABLE:
        if log_func:
            log_func("openpyxl not available; building simple combined workbook without special formatting.")
        return _build_simple_workbook(root_folder, folder_to_case_csvs, log_func)

    # ---------------------------
    # Build scenario DataFrames
    # ---------------------------
    scenario_data = {}  # folder_name -> combined DataFrame

    for folder_name, case_map in folder_to_case_csvs.items():
        dfs = []
        for label in TARGET_PATTERNS:
            csv_path = case_map.get(label)
            if not csv_path:
                continue
            try:
                df = pd.read_csv(csv_path)
                df.insert(0, "CaseType", label)
                dfs.append(df)
            except Exception as e:
                if log_func:
                    log_func(f"  [{folder_name}] WARNING: Failed to read {csv_path}: {e}")

        if dfs:
            scenario_data[folder_name] = pd.concat(dfs, ignore_index=True)

    if not scenario_data:
        if log_func:
            log_func("No scenario data to write into workbook.")
        return None

    # ---------------------------
    # Create formatted workbook
    # ---------------------------
    workbook_path = os.path.join(root_folder, "Combined_ViolationCTG_Comparison.xlsx")

    if log_func:
        log_func(f"\nBuilding FORMATTED combined workbook:\n  {workbook_path}")
        log_func(f"Expandable dropdown grouping is {'ON' if group_details else 'OFF'}.")
        log_func("Sorting Resulting Issues by highest Percent Loading (worst first).")
        log_func("Shifting output down by 1 row (blank Row 1).")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    title_fill = PatternFill(fill_type="solid", fgColor="305496")
    title_font = Font(color="FFFFFF", bold=True, size=12)

    header_fill = PatternFill(fill_type="solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)

    data_font = Font(color="000000")
    data_bold_font = Font(color="000000", bold=True)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Your updated required columns
    required_cols = ["CTGLabel", "LimViolLimit", "LimViolValue", "LimViolPct"]

    for folder_name, df in scenario_data.items():
        sheet_name = (folder_name or "Sheet").strip()[:31] or "Sheet"
        ws = wb.create_sheet(title=sheet_name)

        # Excel outline behavior: summary rows ABOVE details (so dropdown is on the max row)
        ws.sheet_properties.outlinePr.summaryBelow = False

        # Set column widths â€“ contiguous columns B to F
        ws.column_dimensions["B"].width = 55  # Contingency Events
        ws.column_dimensions["C"].width = 55  # Resulting Issue
        ws.column_dimensions["D"].width = 18  # Limit
        ws.column_dimensions["E"].width = 22  # Contingency Value (MVA)
        ws.column_dimensions["F"].width = 18  # Percent Loading

        # Shift everything down by 1 row
        current_row = 2

        for label in TARGET_PATTERNS:
            block_df = df[df["CaseType"] == label].copy()
            if block_df.empty:
                continue

            pretty_name = PRETTY_CASE_NAMES.get(label, label)

            # ===== Title row =====
            # FIX: merge through column F (6) now that we have 5 data cols B-F
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
            c = ws.cell(row=current_row, column=2)
            c.value = pretty_name
            c.fill = title_fill
            c.font = title_font
            c.alignment = center
            for col in range(2, 7):  # B..F
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

            # ===== Header row =====
            # FIX: you were missing a comma after ("D","Limit")
            headers = [
                ("B", "Contingency Events"),
                ("C", "Resulting Issue"),
                ("D", "Limit"),
                ("E", "Contingency Value (MVA)"),
                ("F", "Percent Loading"),
            ]
            for col_letter, text in headers:
                col_idx = ord(col_letter) - ord("A") + 1
                hc = ws.cell(row=current_row, column=col_idx)
                hc.value = text
                hc.fill = header_fill
                hc.font = header_font
                hc.alignment = center
                hc.border = thin_border
            current_row += 1

            # Validate columns
            for col in required_cols:
                if col not in block_df.columns and log_func:
                    log_func(f"  [{folder_name} / {label}] WARNING: column '{col}' missing.")

            has_limviolid = "LimViolID" in block_df.columns

            if group_details and has_limviolid:
                # Numeric percent for sorting
                if "LimViolPct" in block_df.columns:
                    block_df["_pct_num"] = _to_float_series(block_df["LimViolPct"])
                else:
                    block_df["_pct_num"] = pd.Series([float("nan")] * len(block_df), index=block_df.index)

                # Sort within each LimViolID so max is first
                sort_cols = ["LimViolID", "_pct_num"]
                asc = [True, False]
                if "CTGLabel" in block_df.columns:
                    sort_cols.append("CTGLabel")
                    asc.append(True)

                block_df = block_df.sort_values(
                    by=sort_cols, ascending=asc, na_position="last", kind="mergesort"
                )

                # Order groups by their max % (worst first)
                group_max = (
                    block_df.groupby("LimViolID", dropna=False)["_pct_num"]
                    .max()
                    .reset_index()
                    .rename(columns={"_pct_num": "_group_max_pct"})
                )
                group_max["_lim_str"] = group_max["LimViolID"].astype(str)
                group_max = group_max.sort_values(
                    by=["_group_max_pct", "_lim_str"],
                    ascending=[False, True],
                    na_position="last",
                    kind="mergesort",
                )

                ordered_limviolid_values = list(group_max["LimViolID"])

                for limviolid in ordered_limviolid_values:
                    g = block_df[block_df["LimViolID"].eq(limviolid)].copy()
                    if g.empty:
                        continue

                    rows = list(g.itertuples(index=False))
                    if not rows:
                        continue

                    # Summary row (max)
                    r0 = rows[0]

                    ws.cell(row=current_row, column=2).value = getattr(r0, "CTGLabel", "")
                    ws.cell(row=current_row, column=3).value = getattr(r0, "LimViolID", "")
                    ws.cell(row=current_row, column=4).value = getattr(r0, "LimViolLimit", "")
                    ws.cell(row=current_row, column=5).value = getattr(r0, "LimViolValue", "")
                    ws.cell(row=current_row, column=6).value = getattr(r0, "LimViolPct", "")

                    for col in range(2, 7):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = data_bold_font
                        cell.border = thin_border
                        cell.alignment = left_align if col in (2, 3) else center

                    summary_row = current_row
                    current_row += 1

                    # Detail rows (collapsed)
                    detail_start = None
                    detail_end = None

                    for r in rows[1:]:
                        if detail_start is None:
                            detail_start = current_row

                        ws.cell(row=current_row, column=2).value = getattr(r, "CTGLabel", "")
                        ws.cell(row=current_row, column=3).value = ""
                        ws.cell(row=current_row, column=4).value = getattr(r, "LimViolLimit", "")
                        ws.cell(row=current_row, column=5).value = getattr(r, "LimViolValue", "")
                        ws.cell(row=current_row, column=6).value = getattr(r, "LimViolPct", "")

                        for col in range(2, 7):
                            cell = ws.cell(row=current_row, column=col)
                            cell.font = data_font
                            cell.border = thin_border
                            cell.alignment = left_align if col in (2, 3) else center

                        detail_end = current_row
                        current_row += 1

                    if detail_start is not None and detail_end is not None:
                        ws.row_dimensions.group(
                            detail_start,
                            detail_end,
                            outline_level=1,
                            hidden=True,
                        )
                        ws.row_dimensions[summary_row].collapsed = True

            else:
                # No grouping: dump rows
                for _, row in block_df.iterrows():
                    ws.cell(row=current_row, column=2).value = row.get("CTGLabel", "")
                    ws.cell(row=current_row, column=3).value = row.get("LimViolID", "") if has_limviolid else ""
                    ws.cell(row=current_row, column=4).value = row.get("LimViolLimit", "")
                    ws.cell(row=current_row, column=5).value = row.get("LimViolValue", "")
                    ws.cell(row=current_row, column=6).value = row.get("LimViolPct", "")

                    for col in range(2, 7):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = data_font
                        cell.border = thin_border
                        cell.alignment = left_align if col in (2, 3) else center

                    current_row += 1

            # One blank row between blocks
            current_row += 1

    try:
        wb.save(workbook_path)
    except Exception as e:
        if log_func:
            log_func(f"ERROR saving formatted workbook: {e}")
        return None

    if log_func:
        log_func("Formatted combined workbook build complete.")
    return workbook_path