# -*- coding: utf-8 -*-
# gui/trends_view.py

import os
import threading
import queue
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib import rcParams
from itertools import cycle


CASE_TYPES = [
    ("ACCA LongTerm", "acca_longterm"),
    ("ACCA", "acca"),
    ("DCwAC", "dcwac"),
]


def _safe_float(x):
    try:
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip().replace("%", "")
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def _norm(s):
    if s is None:
        return ""
    return str(s).strip()


def _contains(a, b):
    return b.lower() in a.lower()


class TrendsView(ttk.Frame):
    """
    Trends tab:
    - Load a "Combined_Violation..." workbook
    - Treat each sheet as a case
    - For each case type (ACCA LongTerm / ACCA / DCwAC):
        - Extract rows with "Resulting Issue" and "Percent Loading"
        - Build issue -> [loading per sheet]
        - Plot trends across cases
    """

    def __init__(self, master):
        super().__init__(master)

        self._workbook_path = None
        self._sheet_names = []
        self._current_case_type = "acca"

        # Data structure:
        # self._trend_data[case_type]["issues"][issue_key] = list of (sheet_name, loading)
        self._trend_data = {
            ct_key: {"issues": {}, "cases": []} for _, ct_key in CASE_TYPES
        }

        self._ui_queue = queue.Queue()
        self._worker_thread = None
        self._stop_flag = threading.Event()

        self._colors = {}

        self._build_ui()
        self._build_plot()

        # poll UI queue
        self.after(100, self._poll_ui_queue)

    # ---------------- UI ---------------- #

    def _build_ui(self):
        # Top controls
        top = ttk.LabelFrame(self, text="Trends", style="Card.TLabelframe", padding=(10, 8))
        top.pack(fill=tk.X, padx=8, pady=8)

        row1 = ttk.Frame(top)
        row1.pack(fill=tk.X)

        self.btn_open = ttk.Button(row1, text="Open Combined Violation Workbook", command=self._on_open_workbook)
        self.btn_open.pack(side=tk.LEFT)

        self.lbl_loaded = ttk.Label(row1, text="Loaded: No workbook loaded")
        self.lbl_loaded.pack(side=tk.LEFT, padx=12)

        row2 = ttk.Frame(top)
        row2.pack(fill=tk.X, pady=(8, 0))

        ttk.Label(row2, text="Cases (Sheets):").pack(side=tk.LEFT)

        self.case_combo = ttk.Combobox(row2, state="readonly", width=45, values=[])
        self.case_combo.pack(side=tk.LEFT, padx=(8, 12))
        self.case_combo.bind("<<ComboboxSelected>>", self._on_case_selected)

        self.btn_scan = ttk.Button(row2, text="Scan All Sheets", command=self._on_scan_all)
        self.btn_scan.pack(side=tk.LEFT)

        ttk.Label(row2, text="Min %:").pack(side=tk.LEFT, padx=(18, 6))
        self.min_pct_var = tk.StringVar(value="80")
        self.min_pct_entry = ttk.Entry(row2, textvariable=self.min_pct_var, width=6)
        self.min_pct_entry.pack(side=tk.LEFT)

        ttk.Label(row2, text="Top N:").pack(side=tk.LEFT, padx=(18, 6))
        self.topn_var = tk.StringVar(value="8")
        self.topn_entry = ttk.Entry(row2, textvariable=self.topn_var, width=6)
        self.topn_entry.pack(side=tk.LEFT)

        self.btn_plot_top = ttk.Button(row2, text="Plot Top N", command=self._on_plot_top)
        self.btn_plot_top.pack(side=tk.LEFT, padx=(18, 0))

        # Middle body: left issue list + right plot
        mid = ttk.Frame(self, padding=(8, 0))
        mid.pack(fill=tk.BOTH, expand=True)

        left = ttk.LabelFrame(mid, text="Issues", style="Card.TLabelframe", padding=(10, 8))
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 8), pady=8)

        self.issue_tabs = ttk.Notebook(left)
        self.issue_tabs.pack(fill=tk.BOTH, expand=True)

        self._issue_trees = {}
        for display, key in CASE_TYPES:
            frame = ttk.Frame(self.issue_tabs)
            self.issue_tabs.add(frame, text=display)

            tree = ttk.Treeview(frame, columns=("max", "count"), show="headings", height=16)
            tree.heading("max", text="Max %")
            tree.heading("count", text="# Cases")
            tree.column("max", width=70, anchor="e")
            tree.column("count", width=70, anchor="e")

            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            sb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            sb.pack(side=tk.RIGHT, fill=tk.Y)
            tree.configure(yscrollcommand=sb.set)

            tree.bind("<<TreeviewSelect>>", self._on_issue_selected)

            self._issue_trees[key] = tree

        self.issue_tabs.bind("<<NotebookTabChanged>>", self._on_case_type_changed)

        right = ttk.LabelFrame(mid, text="Trend Plot", style="Card.TLabelframe", padding=(10, 8))
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=8)

        self.plot_container = ttk.Frame(right)
        self.plot_container.pack(fill=tk.BOTH, expand=True)

        # Bottom log
        bottom = ttk.LabelFrame(self, text="Log", style="Card.TLabelframe", padding=(10, 8))
        bottom.pack(fill=tk.BOTH, padx=8, pady=(0, 8))

        self.log = tk.Text(bottom, height=6, wrap="none")
        self.log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        sb_y = ttk.Scrollbar(bottom, orient="vertical", command=self.log.yview)
        sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.log.configure(yscrollcommand=sb_y.set)

    def _build_plot(self):
        self.fig = Figure(figsize=(6, 4), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_title("Percent Loading Trend")
        self.ax.set_xlabel("Case (Sheet)")
        self.ax.set_ylabel("Percent Loading (%)")
        self.ax.grid(True)

        self.canvas = FigureCanvasTkAgg(self.fig, master=self.plot_container)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    # ---------------- Logging ---------------- #

    def _log(self, msg):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.update_idletasks()

    # ---------------- Actions ---------------- #

    def _on_open_workbook(self):
        path = filedialog.askopenfilename(
            title="Select Combined Violation Workbook",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not path:
            return

        self._workbook_path = path
        self.lbl_loaded.config(text=f"Loaded: {os.path.basename(path)}")
        self._log(f"Workbook selected: {path}")

        # Load sheet names (quick)
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            self._sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Could not read workbook.\n\n{e}")
            self._sheet_names = []
            return

        self.case_combo["values"] = self._sheet_names
        if self._sheet_names:
            self.case_combo.current(0)
        self._clear_all_data()
        self._log(f"Found {len(self._sheet_names)} sheets.")

    def _on_scan_all(self):
        if not self._workbook_path:
            messagebox.showwarning("No workbook", "Please open a workbook first.")
            return

        if self._worker_thread and self._worker_thread.is_alive():
            messagebox.showinfo("Busy", "Already scanning. Please wait.")
            return

        self._clear_all_data()
        self._stop_flag.clear()
        self._set_busy(True)

        self._worker_thread = threading.Thread(target=self._worker_scan_all, daemon=True)
        self._worker_thread.start()

    def _on_case_selected(self, _evt=None):
        # Optional: scan a single sheet quickly
        if not self._workbook_path:
            return
        sheet = self.case_combo.get().strip()
        if not sheet:
            return

        # Scan just that sheet
        if self._worker_thread and self._worker_thread.is_alive():
            messagebox.showinfo("Busy", "Already scanning. Please wait.")
            return

        self._set_busy(True)
        self._worker_thread = threading.Thread(target=self._worker_scan_one, args=(sheet,), daemon=True)
        self._worker_thread.start()

    def _on_plot_top(self):
        self._refresh_issue_lists()
        self._auto_plot_top()

    def _on_case_type_changed(self, _evt=None):
        tab_idx = self.issue_tabs.index(self.issue_tabs.select())
        self._current_case_type = CASE_TYPES[tab_idx][1]
        self._refresh_issue_lists()

    def _on_issue_selected(self, _evt=None):
        ct = self._current_case_type
        tree = self._issue_trees.get(ct)
        if not tree:
            return
        sel = tree.selection()
        if not sel:
            return
        issue_key = tree.item(sel[0], "text")
        if not issue_key:
            # we stored text differently: use iid as key
            issue_key = sel[0]
        self._plot_specific_issue(ct, issue_key)

    # ---------------- Worker Threads ---------------- #

    def _worker_scan_all(self):
        try:
            path = self._workbook_path
            wb = load_workbook(path, data_only=True, read_only=True)

            for i, sheet_name in enumerate(wb.sheetnames):
                if self._stop_flag.is_set():
                    break

                ws = wb[sheet_name]
                self._ui_queue.put(("log", f"Scanning sheet {i+1}/{len(wb.sheetnames)}: {sheet_name}"))
                parsed = self._parse_sheet(ws, sheet_name)

                # push parsed partial to UI thread
                self._ui_queue.put(("merge", parsed))

            wb.close()
            self._ui_queue.put(("done", None))
        except Exception as e:
            self._ui_queue.put(("error", str(e)))

    def _worker_scan_one(self, sheet_name):
        try:
            path = self._workbook_path
            wb = load_workbook(path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                self._ui_queue.put(("error", f"Sheet not found: {sheet_name}"))
                return

            ws = wb[sheet_name]
            self._ui_queue.put(("log", f"Scanning single sheet: {sheet_name}"))
            parsed = self._parse_sheet(ws, sheet_name)
            wb.close()

            self._ui_queue.put(("merge", parsed))
            self._ui_queue.put(("done", None))
        except Exception as e:
            self._ui_queue.put(("error", str(e)))

    # ---------------- UI Thread Merge ---------------- #

    def _poll_ui_queue(self):
        try:
            while True:
                item = self._ui_queue.get_nowait()
                kind, payload = item

                if kind == "log":
                    self._log(payload)

                elif kind == "merge":
                    self._merge_parsed(payload)

                elif kind == "done":
                    self._log("Scan complete.")
                    self._refresh_issue_lists()
                    self._auto_plot_top()
                    self._set_busy(False)

                elif kind == "error":
                    self._set_busy(False)
                    messagebox.showerror("Error", payload)
                    self._log(f"ERROR: {payload}")

        except queue.Empty:
            pass

        self.after(100, self._poll_ui_queue)

    # ---------------- Parsing Logic ---------------- #

    def _parse_sheet(self, ws, sheet_name):
        """
        Returns dict:
        {
          'acca':   {'case': sheet_name, 'rows': [(issue, pct), ...]},
          'dcwac':  {'case': sheet_name, 'rows': [...]},
          'acca_longterm': ...
        }
        """
        result = {k: {"case": sheet_name, "rows": []} for _, k in CASE_TYPES}

        # We need to find blocks like:
        # header row containing "Resulting Issue" and "Percent Loading"
        # AND a nearby row above that tells case type ("ACCA", "DCwAC", "ACCA LongTerm")
        #
        # The combined violation sheets you showed usually have a big merged header cell
        # with "ACCA" or "DCwAC" above the table.

        max_row = ws.max_row
        max_col = min(ws.max_column, 30)

        # Read a small grid of text for detection
        grid = []
        for r in range(1, max_row + 1):
            row = []
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                row.append(_norm(v))
            grid.append(row)

        # Find table header rows by looking for "Resulting Issue" and "Percent"
        header_rows = []
        for r_idx, row in enumerate(grid, start=1):
            joined = " | ".join(row)
            if _contains(joined, "Resulting Issue") and (_contains(joined, "Percent") or _contains(joined, "Loading")):
                header_rows.append(r_idx)

        # For each detected header row, infer case type from nearby rows above
        for hr in header_rows:
            case_type = self._infer_case_type(grid, hr)
            if not case_type:
                continue

            # Find the columns for "Resulting Issue" and "Percent Loading"
            issue_col, pct_col = self._find_issue_and_pct_cols(grid[hr - 1])

            if issue_col is None or pct_col is None:
                continue

            # Read rows below header until blank-ish
            rows = []
            r = hr + 1
            while r <= max_row:
                issue = _norm(ws.cell(r, issue_col + 1).value)
                pct = _safe_float(ws.cell(r, pct_col + 1).value)

                # Stop if table ended
                if not issue and pct is None:
                    # allow a couple blanks but stop on strong blank
                    # if next few also blank, break
                    blank_run = True
                    for rr in range(r, min(r + 3, max_row) + 1):
                        ii = _norm(ws.cell(rr, issue_col + 1).value)
                        pp = _safe_float(ws.cell(rr, pct_col + 1).value)
                        if ii or (pp is not None):
                            blank_run = False
                            break
                    if blank_run:
                        break

                if issue and pct is not None:
                    rows.append((issue, pct))

                r += 1

            # Append into that case type bucket
            result[case_type]["rows"].extend(rows)

        return result

    def _infer_case_type(self, grid, header_row):
        # Search upward for a row containing one of the case type names
        search_up = 6
        for rr in range(max(1, header_row - search_up), header_row):
            joined = " | ".join(grid[rr - 1])
            j = joined.lower()
            if "acca longterm" in j or "acca long term" in j:
                return "acca_longterm"
            if "dcwac" in j:
                return "dcwac"
            # important: "ACCA" appears in "ACCA LongTerm", so check longterm first
            if "acca" in j:
                return "acca"
        return None

    def _find_issue_and_pct_cols(self, header_cells):
        issue_col = None
        pct_col = None
        for idx, txt in enumerate(header_cells):
            t = txt.lower()
            if "resulting issue" in t:
                issue_col = idx
            if ("percent" in t and "load" in t) or (t.strip() == "percent loading") or ("loading" in t and "percent" in t):
                pct_col = idx
        # If exact name not found, try looser matches
        if pct_col is None:
            for idx, txt in enumerate(header_cells):
                t = txt.lower()
                if "percent" in t:
                    pct_col = idx
                    break
        return issue_col, pct_col

    # ---------------- Data Merge ---------------- #

    def _clear_all_data(self):
        for _, k in CASE_TYPES:
            self._trend_data[k]["issues"].clear()
            self._trend_data[k]["cases"].clear()
        self._colors = {}
        self._clear_issue_lists()
        self._clear_plot()

    def _merge_parsed(self, parsed):
        # parsed[k] => {"case": sheet_name, "rows": [(issue,pct)...]}
        for _, k in CASE_TYPES:
            case_name = parsed[k]["case"]
            rows = parsed[k]["rows"]

            if case_name not in self._trend_data[k]["cases"]:
                self._trend_data[k]["cases"].append(case_name)

            for issue, pct in rows:
                issue_key = self._normalize_issue_key(issue)
                issues = self._trend_data[k]["issues"]
                issues.setdefault(issue_key, [])
                issues[issue_key].append((case_name, pct))

        # update colors after merge
        self._assign_colors()

    def _normalize_issue_key(self, issue):
        # Keep it mostly as-is, but trim whitespace
        return issue.strip()

    # ---------------- Colors (FIXED) ---------------- #

    def _assign_colors(self):
        """
        Stable colors per issue key, based on sorted order.
        Uses matplotlib default color cycle (API-safe).
        """
        ct = self._current_case_type
        issues = self._trend_data.get(ct, {}).get("issues", {})
        keys = sorted(issues.keys())

        default_colors = rcParams["axes.prop_cycle"].by_key().get("color", [])
        if not default_colors:
            default_colors = ["C0", "C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8", "C9"]

        color_cycle = cycle(default_colors)

        self._colors = {}
        for k in keys:
            self._colors[k] = next(color_cycle)

    # ---------------- Issue Lists ---------------- #

    def _clear_issue_lists(self):
        for _, k in CASE_TYPES:
            tree = self._issue_trees.get(k)
            if not tree:
                continue
            for item in tree.get_children():
                tree.delete(item)

    def _refresh_issue_lists(self):
        self._clear_issue_lists()

        min_pct = self._get_min_pct()

        for _, ct_key in CASE_TYPES:
            tree = self._issue_trees[ct_key]
            issues = self._trend_data[ct_key]["issues"]

            # compute max pct and count per issue
            rows = []
            for issue_key, pts in issues.items():
                vals = [p for _, p in pts if p is not None]
                if not vals:
                    continue
                mx = max(vals)
                if mx < min_pct:
                    continue
                rows.append((issue_key, mx, len(set([c for c, _ in pts]))))

            # sort highest to lowest by max
            rows.sort(key=lambda x: x[1], reverse=True)

            for issue_key, mx, cnt in rows:
                # Use iid=issue_key so we can retrieve it cleanly
                tree.insert("", "end", iid=issue_key, values=(f"{mx:.2f}", str(cnt)))

        # refresh colors based on current tab
        self._assign_colors()

    # ---------------- Plotting ---------------- #

    def _clear_plot(self):
        self.ax.clear()
        self.ax.set_title("Percent Loading Trend")
        self.ax.set_xlabel("Case (Sheet)")
        self.ax.set_ylabel("Percent Loading (%)")
        self.ax.grid(True)
        self.canvas.draw_idle()

    def _get_min_pct(self):
        try:
            return float(self.min_pct_var.get().strip())
        except Exception:
            return 80.0

    def _get_topn(self):
        try:
            n = int(float(self.topn_var.get().strip()))
            return max(1, min(n, 50))
        except Exception:
            return 8

    def _auto_plot_top(self):
        ct = self._current_case_type
        issues = self._trend_data[ct]["issues"]
        if not issues:
            self._clear_plot()
            return

        min_pct = self._get_min_pct()
        topn = self._get_topn()

        # rank by max pct
        ranked = []
        for issue_key, pts in issues.items():
            vals = [p for _, p in pts if p is not None]
            if not vals:
                continue
            mx = max(vals)
            if mx < min_pct:
                continue
            ranked.append((issue_key, mx))

        ranked.sort(key=lambda x: x[1], reverse=True)
        ranked = ranked[:topn]

        self._plot_issues(ct, [k for k, _ in ranked], title=f"Top {len(ranked)} Issues Trend ({ct.upper()})")

    def _plot_specific_issue(self, ct, issue_key):
        if issue_key not in self._trend_data[ct]["issues"]:
            return
        self._plot_issues(ct, [issue_key], title=f"Issue Trend ({ct.upper()})")

    def _plot_issues(self, ct, issue_keys, title):
        cases = self._trend_data[ct]["cases"]
        issues = self._trend_data[ct]["issues"]

        if not cases:
            self._clear_plot()
            return

        # Map case name to index
        case_index = {name: i for i, name in enumerate(cases)}

        self.ax.clear()
        self.ax.set_title(title)
        self.ax.set_xlabel("Case (Sheet)")
        self.ax.set_ylabel("Percent Loading (%)")
        self.ax.grid(True)

        # Show fewer x labels if many
        x = list(range(len(cases)))
        self.ax.set_xticks(x)
        if len(cases) <= 12:
            self.ax.set_xticklabels(cases, rotation=35, ha="right", fontsize=8)
        else:
            # sparse labels
            labels = []
            step = max(1, len(cases) // 12)
            for i, name in enumerate(cases):
                labels.append(name if (i % step == 0) else "")
            self.ax.set_xticklabels(labels, rotation=35, ha="right", fontsize=8)

        for issue_key in issue_keys:
            pts = issues.get(issue_key, [])
            # Create series with None for missing cases
            y = [None] * len(cases)
            for c, pct in pts:
                idx = case_index.get(c)
                if idx is not None:
                    y[idx] = pct

            color = self._colors.get(issue_key, None)
            self.ax.plot(x, y, marker="o", linewidth=1.5, label=issue_key, color=color)

        self.ax.legend(loc="best", fontsize=8)
        self.canvas.draw_idle()

    # ---------------- Busy UI ---------------- #

    def _set_busy(self, busy: bool):
        state = "disabled" if busy else "normal"
        try:
            self.btn_open.config(state=state)
            self.btn_scan.config(state=state)
            self.case_combo.config(state="disabled" if busy else "readonly")
            self.btn_plot_top.config(state=state)
        except Exception:
            pass

    # ---------------- Public Hook (optional) ---------------- #

    def set_workbook_path(self, path: str):
        """
        Optional external hook: if another tab loads a workbook and wants to share it.
        """
        if not path or not os.path.exists(path):
            return
        self._workbook_path = path
        self.lbl_loaded.config(text=f"Loaded: {os.path.basename(path)}")
        self._log(f"Workbook path set externally: {path}")

        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            self._sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            self._log(f"ERROR reading workbook: {e}")
            self._sheet_names = []
            return

        self.case_combo["values"] = self._sheet_names
        if self._sheet_names:
            self.case_combo.current(0)
        self._clear_all_data()