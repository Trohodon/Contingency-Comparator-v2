# tab_compare.py
# GUI tab for comparing sheets and building queued workbooks

from __future__ import annotations

import os
import threading
import queue
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from comparator import build_case_type_comparison, build_batch_comparison_workbook


class TabCompare(ttk.Frame):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)

        self.workbook_path = tk.StringVar(value="")
        self.left_sheet = tk.StringVar(value="")
        self.right_sheet = tk.StringVar(value="")

        self.left_name = tk.StringVar(value="Left")
        self.right_name = tk.StringVar(value="Right")

        self.expandable_issue_view = tk.BooleanVar(value=True)

        self._log_q: "queue.Queue[str]" = queue.Queue()
        self._worker_thread: threading.Thread | None = None

        # queued batch tasks
        self._queued_tasks = []

        self._build_ui()

        # Start log pump
        self.after(100, self._pump_log)

    # -------------------------
    # UI
    # -------------------------

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=10)

        ttk.Label(top, text="Workbook:").grid(row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.workbook_path, width=65).grid(row=0, column=1, sticky="we", padx=5)
        ttk.Button(top, text="Browse", command=self._browse_workbook).grid(row=0, column=2, sticky="e")

        top.columnconfigure(1, weight=1)

        # Compare frame
        cmp_frame = ttk.LabelFrame(self, text="Compare")
        cmp_frame.pack(fill="x", padx=10, pady=(0, 10))

        ttk.Label(cmp_frame, text="Left sheet:").grid(row=0, column=0, sticky="w", padx=5, pady=4)
        self.left_cb = ttk.Combobox(cmp_frame, textvariable=self.left_sheet, width=45, state="readonly")
        self.left_cb.grid(row=0, column=1, sticky="w", padx=5, pady=4)

        ttk.Label(cmp_frame, text="Right sheet:").grid(row=1, column=0, sticky="w", padx=5, pady=4)
        self.right_cb = ttk.Combobox(cmp_frame, textvariable=self.right_sheet, width=45, state="readonly")
        self.right_cb.grid(row=1, column=1, sticky="w", padx=5, pady=4)

        ttk.Label(cmp_frame, text="Left name:").grid(row=0, column=2, sticky="w", padx=5, pady=4)
        ttk.Entry(cmp_frame, textvariable=self.left_name, width=18).grid(row=0, column=3, sticky="w", padx=5, pady=4)

        ttk.Label(cmp_frame, text="Right name:").grid(row=1, column=2, sticky="w", padx=5, pady=4)
        ttk.Entry(cmp_frame, textvariable=self.right_name, width=18).grid(row=1, column=3, sticky="w", padx=5, pady=4)

        ttk.Checkbutton(
            cmp_frame,
            text="Expandable issue view (Excel +/-)",
            variable=self.expandable_issue_view,
        ).grid(row=2, column=1, sticky="w", padx=5, pady=4)

        ttk.Button(cmp_frame, text="Compare (Live View)", command=self._compare_live).grid(
            row=0, column=5, sticky="e", padx=5, pady=4
        )
        ttk.Button(cmp_frame, text="Queue Pair", command=self._queue_pair).grid(
            row=1, column=5, sticky="e", padx=5, pady=4
        )

        build_btn = ttk.Button(cmp_frame, text="Build Queued Workbook", command=self._build_queued_workbook)
        build_btn.grid(row=2, column=5, sticky="e", padx=5, pady=4)
        note_lbl = ttk.Label(
            cmp_frame,
            text="Note: Build queued workbook also adds a final 'Compare' sheet (straight comparison).",
            wraplength=380,
            justify="left",
        )
        note_lbl.grid(row=3, column=5, sticky="nw", padx=(5, 5), pady=(0, 6))

        # Live view table
        table_frame = ttk.Frame(self)
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ("CaseType", "Contingency", "Issue", "LeftPct", "RightPct", "Delta")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=12)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=140 if c in ("Contingency", "Issue") else 100, anchor="w")
        self.tree.pack(side="left", fill="both", expand=True)

        scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scroll.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scroll.set)

        # Log
        log_frame = ttk.LabelFrame(self, text="Log")
        log_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.log = tk.Text(log_frame, height=10, wrap="word")
        self.log.pack(fill="both", expand=True)

    # -------------------------
    # Actions
    # -------------------------

    def _browse_workbook(self):
        path = filedialog.askopenfilename(
            title="Select workbook",
            filetypes=[("Excel workbooks", "*.xlsx")],
        )
        if not path:
            return
        self.workbook_path.set(path)
        self._refresh_sheet_lists()

    def _refresh_sheet_lists(self):
        # Lazy import openpyxl here (keeps app fast on startup)
        try:
            from openpyxl import load_workbook
        except Exception as e:
            messagebox.showerror("Error", f"openpyxl not available: {e}")
            return

        path = self.workbook_path.get().strip()
        if not os.path.isfile(path):
            return

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read workbook: {e}")
            return

        self.left_cb["values"] = sheets
        self.right_cb["values"] = sheets

        if sheets:
            if not self.left_sheet.get():
                self.left_sheet.set(sheets[0])
            if not self.right_sheet.get() and len(sheets) > 1:
                self.right_sheet.set(sheets[1] if sheets[1] != self.left_sheet.get() else sheets[0])

    def _compare_live(self):
        path = self.workbook_path.get().strip()
        if not os.path.isfile(path):
            messagebox.showwarning("Missing workbook", "Please select a workbook first.")
            return

        left = self.left_sheet.get().strip()
        right = self.right_sheet.get().strip()
        if not left or not right:
            messagebox.showwarning("Missing selection", "Select both left and right sheets.")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            if left not in wb.sheetnames or right not in wb.sheetnames:
                messagebox.showerror("Error", "Selected sheets not found in workbook.")
                return

            df = build_case_type_comparison(
                wb[left],
                wb[right],
                left_name=self.left_name.get().strip() or "Left",
                right_name=self.right_name.get().strip() or "Right",
            )
        except Exception as e:
            messagebox.showerror("Error", f"Compare failed: {e}")
            return

        # populate tree
        for item in self.tree.get_children():
            self.tree.delete(item)

        if df.empty:
            return

        # show top N rows
        for _, r in df.head(500).iterrows():
            self.tree.insert(
                "",
                "end",
                values=(
                    r.get("CaseType", ""),
                    r.get("CTGLabel", ""),
                    r.get("LimViolID", ""),
                    r.get("LimViolPct (Left)", ""),
                    r.get("LimViolPct (Right)", ""),
                    r.get("Delta", ""),
                ),
            )

    def _queue_pair(self):
        path = self.workbook_path.get().strip()
        if not os.path.isfile(path):
            messagebox.showwarning("Missing workbook", "Please select a workbook first.")
            return

        left = self.left_sheet.get().strip()
        right = self.right_sheet.get().strip()
        if not left or not right:
            messagebox.showwarning("Missing selection", "Select both left and right sheets.")
            return

        left_name = self.left_name.get().strip() or "Left"
        right_name = self.right_name.get().strip() or "Right"

        # Sheet naming uses " vs " (no numbering)
        sheet_name = f"{left_name} vs {right_name}"

        self._queued_tasks.append(
            {
                "LeftSheet": left,
                "RightSheet": right,
                "LeftName": left_name,
                "RightName": right_name,
                "SheetName": sheet_name,
            }
        )
        self._log(f"Queued: {sheet_name}")

    def _build_queued_workbook(self):
        if not self._queued_tasks:
            messagebox.showinfo("Nothing queued", "Queue at least one pair before building.")
            return

        out = filedialog.asksaveasfilename(
            title="Save output workbook",
            defaultextension=".xlsx",
            filetypes=[("Excel workbooks", "*.xlsx")],
        )
        if not out:
            return

        src_path = self.workbook_path.get().strip()
        if not os.path.isfile(src_path):
            messagebox.showwarning("Missing workbook", "Please select a workbook first.")
            return

        if self._worker_thread and self._worker_thread.is_alive():
            messagebox.showinfo("Busy", "A build is already running.")
            return

        tasks = list(self._queued_tasks)
        expandable = self.expandable_issue_view.get()

        self._log(f"Building workbook: {out}")
        self._log("This may take a bit for large batches...")

        def worker():
            try:
                build_batch_comparison_workbook(
                    tasks,
                    out,
                    src_workbook=src_path,
                    expandable_issue_view=expandable,
                    log_func=self._log,
                )
                self._log("Done.")
            except Exception as e:
                self._log(f"ERROR: {e}")
                messagebox.showerror("Error", f"Failed to build workbook: {e}")

        self._worker_thread = threading.Thread(target=worker, daemon=True)
        self._worker_thread.start()

    # -------------------------
    # Logging
    # -------------------------

    def _log(self, msg: str):
        try:
            self._log_q.put_nowait(msg)
        except Exception:
            pass

    def _pump_log(self):
        try:
            while True:
                msg = self._log_q.get_nowait()
                self.log.insert("end", msg + "\n")
                self.log.see("end")
        except queue.Empty:
            pass
        self.after(100, self._pump_log)