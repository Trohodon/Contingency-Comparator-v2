import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from core.case_finder import _classify_case
from core.case_types import CASE_TYPE_DEFINITIONS
from core.comparator import (
    build_all_case_type_comparisons,
    build_batch_comparison_workbook,
)


def _write_scenario_sheet(ws, offset):
    row = 2
    for canonical, pretty in CASE_TYPE_DEFINITIONS:
        ws.cell(row=row, column=2).value = pretty
        row += 1
        headers = [
            "Contingency Events",
            "Resulting Issue",
            "Limit",
            "Contingency Value (MVA)",
            "Percent Loading",
        ]
        for col, header in enumerate(headers, start=2):
            ws.cell(row=row, column=col).value = header
        row += 1

        ws.cell(row=row, column=2).value = f"{canonical} CTG"
        ws.cell(row=row, column=3).value = f"{canonical} Issue"
        ws.cell(row=row, column=4).value = 100
        ws.cell(row=row, column=5).value = 90 + offset
        ws.cell(row=row, column=6).value = 90 + offset
        row += 1

        if canonical == "AUXapplied":
            ws.cell(row=row, column=2).value = f"{canonical} CTG"
            ws.cell(row=row, column=3).value = f"{canonical} Issue"
            ws.cell(row=row, column=4).value = 100
            ws.cell(row=row, column=5).value = 95 + offset
            ws.cell(row=row, column=6).value = f"{95 + offset}%"
            row += 1

        row += 1


class CaseTypeIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.source_path = os.path.join(self.temp_dir.name, "source.xlsx")

        wb = Workbook()
        left = wb.active
        left.title = "Left"
        right = wb.create_sheet("Right")
        _write_scenario_sheet(left, 0)
        _write_scenario_sheet(right, 5)
        wb.save(self.source_path)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_auxapplied_filename_recognition_is_case_insensitive(self):
        self.assertEqual(_classify_case("Study_AUXapplied_Final.pwb"), "AUXapplied")
        self.assertEqual(_classify_case("study_auxapplied_final.PWB"), "AUXapplied")

    def test_all_case_types_compare_and_duplicate_keys_use_maximum(self):
        comparisons = build_all_case_type_comparisons(
            self.source_path, "Left", "Right"
        )

        self.assertEqual(set(comparisons), {c for c, _p in CASE_TYPE_DEFINITIONS})
        aux = comparisons["AUXapplied"]
        self.assertEqual(len(aux), 1)
        self.assertEqual(aux.iloc[0]["LeftPct"], 95)
        self.assertEqual(aux.iloc[0]["RightPct"], 100)
        self.assertEqual(aux.iloc[0]["DeltaPct"], 5)

    def test_batch_and_straight_outputs_include_auxapplied(self):
        output_path = os.path.join(self.temp_dir.name, "comparison.xlsx")
        build_batch_comparison_workbook(
            src_workbook=self.source_path,
            pairs=[("Left", "Right")],
            threshold=80,
            output_path=output_path,
        )

        wb = load_workbook(output_path, read_only=True, data_only=True)
        self.assertIn("Left vs Right", wb.sheetnames)
        self.assertIn("Straight Comparison", wb.sheetnames)

        pair_titles = [
            row[0]
            for row in wb["Left vs Right"].iter_rows(
                min_col=2, max_col=2, values_only=True
            )
            if row[0]
        ]
        straight_titles = [
            row[0]
            for row in wb["Straight Comparison"].iter_rows(
                min_col=2, max_col=2, values_only=True
            )
            if row[0]
        ]
        self.assertIn("AUXapplied", pair_titles)
        self.assertIn("AUXapplied", straight_titles)
        wb.close()

    def test_high_threshold_still_creates_valid_workbook(self):
        output_path = os.path.join(self.temp_dir.name, "empty.xlsx")
        build_batch_comparison_workbook(
            src_workbook=self.source_path,
            pairs=[("Left", "Right")],
            threshold=500,
            output_path=output_path,
        )

        wb = load_workbook(output_path, read_only=True, data_only=True)
        self.assertEqual(
            wb["Left vs Right"].cell(row=2, column=2).value,
            "No rows above threshold.",
        )
        wb.close()


if __name__ == "__main__":
    unittest.main()
